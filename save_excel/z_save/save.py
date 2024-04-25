
import openpyxl
import pandas as pd


path = r'C:\Users\86185\Desktop\20230619.xlsx'
path1 = r'C:\Users\86185\Desktop\sl.xlsx'
df = pd.read_excel(path, engine='openpyxl', sheet_name='Sheet1')
def get_all_Data():
    """
    @return:返回G列的所有数据
    """

    col_data = df.iloc[:, 6]#查找G列数据
    all_data= []
    for col in col_data:
        all_data.append(col)
    return all_data
def get_GN_data():
    """

    @return:返回的是GN=
    """
    GN_DATA = []
    all_data = get_all_Data()
    for i in all_data:
        try:
            GN = (i.split("GN=")[1].split("PE")[0])
            GN_DATA.append(GN)
        except:
            GN_DATA.append("这行没有")

    return GN_DATA

def get_frist_data():
    frist_DATA = []
    all_data = get_all_Data()
    true_data = []
    """
    不好分割i先分一半，s在获取最终的值
    """
    for i in all_data:
        try:
            GN = (i.split("sp|")[1].split("_")[0])
            frist_DATA.append(GN)
        except:
            frist_DATA.append("这行没有")

    for s in frist_DATA:
        try:
            GN = (s.split("|")[1].split(" ")[0])
            true_data.append(GN)
        except:
            true_data.append("这行没有")
    return true_data

def save_data():
    # 打开Excel文件
    workbook = openpyxl.load_workbook(path)
    # 获取第一个工作表
    worksheet = workbook.active
    # 在第一列添加表头
    worksheet.insert_cols(1)
    worksheet['Y1'] = 'frist_data'
    worksheet['Z1'] = 'GN_data'
    # 在第一列添加数据
    frist_data = get_frist_data()
    gn_LIST = get_GN_data()

    for row, value in enumerate(frist_data):
        worksheet.cell(row=row + 2, column=25, value=value)
    # 保存Excel文件
    for row, value in enumerate(gn_LIST):
        worksheet.cell(row=row + 2, column=26, value=value)
    workbook.save(path1)




save_data()

